"""Hazard profiling pipeline: FEMA NRI + USFS Wildfire Risk ingestion.

Builds per-Tribe hazard profiles by:
  1. Parsing FEMA National Risk Index (NRI) county-level CSV data
  2. Bridging NRI counties to Tribal areas via AIANNH crosswalk
  3. Parsing USFS Wildfire Risk to Communities XLSX data
  4. Aggregating and caching per-Tribe hazard profiles as JSON

The NRI does NOT have Tribal areas as a primary geographic unit. FEMA provides
a "Tribal County Relational Database" CSV that maps AIANNH boundaries to
overlapping NRI counties. The existing aiannh_tribe_crosswalk.json bridges
Census AIANNH GEOIDs to EPA tribe_ids.

Usage:
    builder = HazardProfileBuilder(config)
    count = builder.build_all_profiles()
    print(f"Built {count} hazard profiles")
"""

import csv
import json
import logging
import math
from datetime import datetime, timezone
from pathlib import Path

from src.packets._geo_common import (
    atomic_write_json,
    atomic_write_text,
    load_aiannh_crosswalk,
    load_area_weights,
)
from src.paths import (
    AIANNH_CROSSWALK_PATH,
    HAZARD_PROFILES_DIR,
    NRI_DIR,
    OUTPUTS_DIR,
    TRIBAL_COUNTY_WEIGHTS_PATH,
    USFS_DIR,
    resolve_path,
)

logger = logging.getLogger(__name__)

# All 18 NRI hazard type codes and their display names
NRI_HAZARD_CODES = {
    "AVLN": "Avalanche",
    "CFLD": "Coastal Flooding",
    "CWAV": "Cold Wave",
    "DRGT": "Drought",
    "ERQK": "Earthquake",
    "HAIL": "Hail",
    "HWAV": "Heat Wave",
    "HRCN": "Hurricane",
    "ISTM": "Ice Storm",
    "IFLD": "Inland Flooding",
    "LNDS": "Landslide",
    "LTNG": "Lightning",
    "SWND": "Strong Wind",
    "TRND": "Tornado",
    "TSUN": "Tsunami",
    "VLCN": "Volcanic Activity",
    "WFIR": "Wildfire",
    "WNTW": "Winter Weather",
}

# US state abbreviation to FIPS prefix mapping for state-level fallback
_STATE_FIPS_PREFIX = {
    "AL": "01", "AK": "02", "AZ": "04", "AR": "05", "CA": "06",
    "CO": "08", "CT": "09", "DE": "10", "DC": "11", "FL": "12",
    "GA": "13", "HI": "15", "ID": "16", "IL": "17", "IN": "18",
    "IA": "19", "KS": "20", "KY": "21", "LA": "22", "ME": "23",
    "MD": "24", "MA": "25", "MI": "26", "MN": "27", "MS": "28",
    "MO": "29", "MT": "30", "NE": "31", "NV": "32", "NH": "33",
    "NJ": "34", "NM": "35", "NY": "36", "NC": "37", "ND": "38",
    "OH": "39", "OK": "40", "OR": "41", "PA": "42", "RI": "44",
    "SC": "45", "SD": "46", "TN": "47", "TX": "48", "UT": "49",
    "VT": "50", "VA": "51", "WA": "53", "WV": "54", "WI": "55",
    "WY": "56",
}


def _safe_float(value, default: float | None = 0.0) -> float | None:
    """Convert a value to float, returning default for empty/None/non-numeric.

    Args:
        value: The value to convert. May be str, None, int, float, or empty.
        default: Fallback value if conversion fails. When default=None, None
            may be returned to signal absence of a value (caller must handle).

    Returns:
        Float value, or default (which may be None when explicitly passed).
    """
    if value is None:
        return default
    if isinstance(value, (int, float)):
        result = float(value)
        if math.isnan(result) or math.isinf(result):
            return default
        return result
    s = str(value).strip()
    if not s:
        return default
    try:
        result = float(s)
        if math.isnan(result) or math.isinf(result):
            return default
        return result
    except (ValueError, TypeError):
        return default


def score_to_rating(score: float) -> str:
    """Convert 0-100 percentile score to NRI rating string.

    Uses quintile breakpoints as approximation of FEMA's methodology.
    SOVI/RESL officially use quintiles. Risk/EAL officially use k-means
    but quintiles are a defensible approximation for area-weighted derived scores.

    Args:
        score: Percentile score in range 0-100.

    Returns:
        NRI rating string from "Very Low" to "Very High".
    """
    if score >= 80:
        return "Very High"
    elif score >= 60:
        return "Relatively High"
    elif score >= 40:
        return "Relatively Moderate"
    elif score >= 20:
        return "Relatively Low"
    else:
        return "Very Low"


class HazardProfileBuilder:
    """Builds per-Tribe hazard profiles from FEMA NRI and USFS data.

    Loads NRI county-level CSV data, bridges to Tribal areas via the
    AIANNH crosswalk, adds USFS wildfire metrics, and writes per-Tribe
    JSON cache files to data/hazard_profiles/.

    Attributes:
        crosswalk_path: Path to aiannh_tribe_crosswalk.json.
        nri_dir: Path to directory containing NRI CSV files.
        usfs_dir: Path to directory containing USFS XLSX files.
        cache_dir: Path to output directory for per-Tribe cache files.
    """

    def __init__(self, config: dict) -> None:
        """Initialize the builder.

        Args:
            config: Application configuration dict. Reads paths from
                config["packets"]["hazards"][*], falling back to defaults.
        """
        from src.packets.registry import TribalRegistry

        packets_cfg = config.get("packets", {})
        hazards_cfg = packets_cfg.get("hazards", {})

        raw = hazards_cfg.get("crosswalk_path")
        self.crosswalk_path = resolve_path(raw) if raw else AIANNH_CROSSWALK_PATH

        raw = hazards_cfg.get("nri_dir")
        self.nri_dir = resolve_path(raw) if raw else NRI_DIR

        raw = hazards_cfg.get("usfs_dir")
        self.usfs_dir = resolve_path(raw) if raw else USFS_DIR

        raw = hazards_cfg.get("cache_dir")
        self.cache_dir = resolve_path(raw) if raw else HAZARD_PROFILES_DIR

        # Load registry for tribe enumeration
        self._registry = TribalRegistry(config)

        # Load crosswalk: AIANNH GEOID -> tribe_id
        self._crosswalk, self._tribe_to_geoids = load_aiannh_crosswalk(
            self.crosswalk_path, label="hazard",
        )

        # Load pre-computed area-weighted AIANNH-to-county crosswalk
        self._area_weights = load_area_weights(
            TRIBAL_COUNTY_WEIGHTS_PATH, label="hazard",
        )

        # NRI dataset version -- detected dynamically in _load_nri_county_data()
        self._nri_version: str = "NRI_v1.20"

    @staticmethod
    def _detect_nri_version(csv_path: Path) -> str:
        """Detect NRI dataset version from file path or directory structure.

        FEMA distributes NRI data from versioned URL paths like
        ``/nri/v120/NRI_Table_Counties.zip``. The download script
        saves to ``data/nri/``, so we check parent directory names
        and ZIP filenames for version patterns.

        Args:
            csv_path: Path to the NRI county CSV file.

        Returns:
            Version string like "NRI_v1.20". Falls back to "NRI_v1.20"
            if detection fails.
        """
        import re

        default_version = "NRI_v1.20"

        # Check for version in sibling ZIP filenames (e.g., NRI_v1.20_counties.zip)
        nri_dir = csv_path.parent
        try:
            siblings = list(nri_dir.iterdir())
        except OSError:
            siblings = []  # Fall through to default version
        for sibling in siblings:
            name_lower = sibling.name.lower()
            # Look for patterns like "v1.20", "v120", "v1_20"
            match = re.search(r"v(\d+)[._]?(\d+)", name_lower)
            if match:
                major = match.group(1)
                minor = match.group(2)
                if len(major) >= 2 and len(minor) >= 2:
                    # "v120" -> "1.20"
                    version = f"NRI_v{major[0]}.{major[1:]}"
                else:
                    version = f"NRI_v{major}.{minor}"
                logger.debug("Detected NRI version %s from %s", version, sibling.name)
                return version

        logger.debug(
            "NRI version not detected from %s, using default %s",
            nri_dir, default_version,
        )
        return default_version

    def _load_nri_county_data(self) -> dict[str, dict]:
        """Parse the NRI county-level CSV file.

        Reads data/nri/NRI_Table_Counties.csv and extracts composite
        risk metrics plus per-hazard scores for all 18 hazard types.

        Returns:
            Dict keyed by STCOFIPS (state+county FIPS) -> metrics dict.
            Empty dict if the file does not exist.
        """
        csv_path = self.nri_dir / "NRI_Table_Counties.csv"
        if not csv_path.exists():
            logger.warning(
                "NRI county CSV not found at %s -- "
                "NRI hazard data will be unavailable. "
                "Download from https://hazards.fema.gov/nri/data-resources",
                csv_path,
            )
            return {}

        county_data: dict[str, dict] = {}
        logger.info("Loading NRI county data from %s", csv_path)

        # Detect NRI version from directory structure or filename
        # FEMA distributes as NRI_Table_Counties.zip from versioned URL paths
        # e.g. .../nri/v120/NRI_Table_Counties.zip -> version "NRI_v1.20"
        self._nri_version = self._detect_nri_version(csv_path)

        with open(csv_path, encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            row_count = 0
            for row in reader:
                fips = row.get("STCOFIPS", "").strip()
                fips = fips.zfill(5) if fips else fips
                if not fips:
                    continue

                record: dict = {
                    "fips": fips,
                    "county": row.get("COUNTY", ""),
                    "state": row.get("STATE", ""),
                    # Composite risk metrics
                    "risk_score": _safe_float(row.get("RISK_SCORE")),
                    "risk_rating": row.get("RISK_RATNG", "").strip(),
                    "risk_value": _safe_float(row.get("RISK_VALUE")),
                    "eal_total": _safe_float(row.get("EAL_VALT")),
                    "eal_score": _safe_float(row.get("EAL_SCORE")),
                    "eal_rating": row.get("EAL_RATNG", "").strip(),
                    "sovi_score": _safe_float(row.get("SOVI_SCORE")),
                    "sovi_rating": row.get("SOVI_RATNG", "").strip(),
                    "resl_score": _safe_float(row.get("RESL_SCORE")),
                    "resl_rating": row.get("RESL_RATNG", "").strip(),
                    # Per-hazard metrics
                    "hazards": {},
                }

                for code in NRI_HAZARD_CODES:
                    record["hazards"][code] = {
                        "risk_score": _safe_float(row.get(f"{code}_RISKS")),
                        "risk_rating": row.get(f"{code}_RISKR", "").strip(),
                        "eal_total": _safe_float(row.get(f"{code}_EALT")),
                        "annualized_freq": _safe_float(row.get(f"{code}_AFREQ")),
                        "num_events": _safe_float(row.get(f"{code}_EVNTS")),
                    }

                county_data[fips] = record
                row_count += 1

        logger.info("Loaded NRI data for %d counties", row_count)
        return county_data

    def _load_nri_tribal_relational(self) -> dict[str, list[str]]:
        """Parse the NRI Tribal County Relational CSV.

        Maps AIANNH area GEOIDs to lists of overlapping county FIPS codes.
        Searches data/nri/ for CSV files with 'Tribal' and 'Relational'
        in the filename.

        Returns:
            Dict mapping AIANNH GEOID -> list of county FIPS codes.
            Empty dict if the relational CSV is not found.
        """
        # Search for the tribal relational CSV by pattern
        nri_path = self.nri_dir
        if not nri_path.exists():
            logger.warning(
                "NRI data directory %s does not exist -- "
                "tribal-county relational data unavailable",
                nri_path,
            )
            return {}

        candidates = []
        for csv_file in nri_path.iterdir():
            name_lower = csv_file.name.lower()
            if (csv_file.suffix.lower() == ".csv"
                    and "tribal" in name_lower
                    and "relational" in name_lower):
                candidates.append(csv_file)

        if not candidates:
            logger.warning(
                "No NRI Tribal County Relational CSV found in %s -- "
                "will fall back to state-level county matching",
                nri_path,
            )
            return {}

        csv_path = candidates[0]
        if len(candidates) > 1:
            logger.info(
                "Found %d tribal relational CSVs, using: %s",
                len(candidates), csv_path.name,
            )

        logger.info("Loading NRI tribal relational data from %s", csv_path)
        tribal_counties: dict[str, list[str]] = {}

        with open(csv_path, encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            headers = reader.fieldnames or []
            logger.info("Tribal relational CSV headers: %s", headers)

            # Discover column names -- look for AIANNH GEOID and county FIPS
            geoid_col = None
            fips_col = None
            for h in headers:
                h_upper = h.upper()
                if geoid_col is None and any(
                    kw in h_upper for kw in ["AIESSION", "AESSION", "GEOID", "AIANNH", "TRIBAL_ID"]
                ):
                    geoid_col = h
                if fips_col is None and any(
                    kw in h_upper for kw in ["STCOFIPS", "COUNTYFIPS", "COUNTY_FIPS", "CNTY_FIPS"]
                ):
                    fips_col = h

            if geoid_col is None or fips_col is None:
                # Broader search: any column with 'geoid' or 'fips'
                for h in headers:
                    h_lower = h.lower()
                    if geoid_col is None and "geoid" in h_lower:
                        geoid_col = h
                    if fips_col is None and "fips" in h_lower and "state" not in h_lower:
                        fips_col = h

            if geoid_col is None or fips_col is None:
                logger.error(
                    "Cannot identify GEOID and FIPS columns in %s. "
                    "Headers: %s. Tried GEOID/AIANNH and STCOFIPS/COUNTYFIPS patterns.",
                    csv_path, headers,
                )
                return {}

            logger.info(
                "Using columns: GEOID='%s', FIPS='%s'",
                geoid_col, fips_col,
            )

            for row in reader:
                geoid = row.get(geoid_col, "").strip()
                fips = row.get(fips_col, "").strip()
                if geoid and fips:
                    if geoid not in tribal_counties:
                        tribal_counties[geoid] = []
                    if fips not in tribal_counties[geoid]:
                        tribal_counties[geoid].append(fips)

        logger.info(
            "Loaded tribal-county relations: %d AIANNH areas -> %d total county links",
            len(tribal_counties),
            sum(len(v) for v in tribal_counties.values()),
        )
        return tribal_counties

    def _get_state_county_fips(
        self, states: list[str], county_data: dict[str, dict]
    ) -> list[str]:
        """Fallback: get all county FIPS codes for given states.

        Used when the NRI tribal relational CSV is unavailable. Less precise
        than AIANNH-to-county matching but ensures every Tribe gets some data.

        Args:
            states: List of 2-letter state abbreviations.
            county_data: Full NRI county data dict keyed by STCOFIPS.

        Returns:
            List of county FIPS codes whose state prefix matches.
        """
        state_prefixes = set()
        for st in states:
            prefix = _STATE_FIPS_PREFIX.get(st.upper())
            if prefix:
                state_prefixes.add(prefix)

        if not state_prefixes:
            return []

        return [
            fips for fips in county_data
            if fips[:2] in state_prefixes
        ]

    def _build_tribe_nri_profile(
        self,
        tribe_id: str,
        tribe: dict,
        county_data: dict[str, dict],
        tribal_counties: dict[str, list[str]],
    ) -> dict:
        """Aggregate NRI data for a single Tribe using area-weighted averaging.

        Resolution chain:
          1. tribe_id -> AIANNH GEOIDs (from crosswalk)
          2. AIANNH GEOIDs -> county FIPS + area weights (from area crosswalk)
          3. Fallback: AIANNH GEOIDs -> county FIPS (from tribal relational CSV)
          4. Fallback: state-level county matching (equal weights)

        Aggregation strategy (area-weighted):
          - Percentile scores (risk_score, eal_score, sovi_score, resl_score):
            weighted AVERAGE proportional to geographic overlap area
          - Dollar amounts (eal_total): weighted average (proportional attribution)
          - Rating strings: re-derived from weighted scores via score_to_rating()
          - Non-zero hazards only: zero-score hazard types are omitted
          - Top 5 hazards: ranked by risk_score descending

        Args:
            tribe_id: EPA tribe identifier.
            tribe: Tribe dict from registry.
            county_data: Full NRI county data dict.
            tribal_counties: AIANNH GEOID -> county FIPS mapping.

        Returns:
            NRI profile dict with composite scores, top hazards, and all hazards.
        """
        if not county_data:
            return self._empty_nri_profile("NRI county data not loaded")

        # Step 1: Find AIANNH GEOIDs for this Tribe
        geoids = self._tribe_to_geoids.get(tribe_id, [])

        # Step 2: Collect county weights from area-weighted crosswalk
        county_weights: dict[str, float] = {}  # fips -> weight
        county_fips_set: set[str] = set()

        if geoids and self._area_weights:
            for geoid in geoids:
                entries = self._area_weights.get(geoid, [])
                for entry in entries:
                    fips = entry["county_fips"]
                    weight = entry["weight"]
                    # If same county appears via multiple GEOIDs, sum the weights
                    county_weights[fips] = county_weights.get(fips, 0.0) + weight
            county_fips_set = set(county_weights.keys())

        # Fallback: relational CSV with equal weights
        if not county_weights and geoids and tribal_counties:
            for geoid in geoids:
                for fips in tribal_counties.get(geoid, []):
                    county_fips_set.add(fips)
            if county_fips_set:
                equal_w = 1.0 / len(county_fips_set)
                county_weights = {fips: equal_w for fips in county_fips_set}

        # Last resort: state-level fallback with equal weights
        if not county_fips_set:
            states = tribe.get("states", [])
            if states:
                state_fips = self._get_state_county_fips(states, county_data)
                county_fips_set.update(state_fips)
                if county_fips_set:
                    tribe_id = tribe.get("epa_id", tribe.get("id", "unknown"))
                    logger.info(
                        "Tribe %s: using state-level NRI fallback (%d counties from %d states)",
                        tribe_id, len(county_fips_set), len(states),
                    )
                    equal_w = 1.0 / len(county_fips_set)
                    county_weights = {fips: equal_w for fips in county_fips_set}

        if not county_fips_set:
            return self._empty_nri_profile(
                "No county matches found (no crosswalk entry or state data)"
            )

        # Step 3: Collect matching county NRI records
        matched_counties = []
        for fips in county_fips_set:
            if fips in county_data:
                matched_counties.append(county_data[fips])

        if not matched_counties:
            return self._empty_nri_profile(
                f"Counties identified ({len(county_fips_set)}) but none found in NRI data"
            )

        # Step 4: Area-weighted composite scores
        # Normalize weights to matched counties only
        total_w = sum(county_weights.get(c["fips"], 0.0) for c in matched_counties)
        if total_w == 0:
            total_w = len(matched_counties)  # Equal weight fallback
            norm_weights = {c["fips"]: 1.0 / total_w for c in matched_counties}
        else:
            norm_weights = {
                c["fips"]: county_weights.get(c["fips"], 0.0) / total_w
                for c in matched_counties
            }

        # Percentile scores: weighted AVERAGE
        composite = {
            "risk_score": round(
                sum(c["risk_score"] * norm_weights[c["fips"]] for c in matched_counties), 2
            ),
            "eal_score": round(
                sum(c["eal_score"] * norm_weights[c["fips"]] for c in matched_counties), 2
            ),
            "sovi_score": round(
                sum(c["sovi_score"] * norm_weights[c["fips"]] for c in matched_counties), 2
            ),
            "resl_score": round(
                sum(c["resl_score"] * norm_weights[c["fips"]] for c in matched_counties), 2
            ),
            # Dollar amounts: weighted average (proportional attribution)
            "eal_total": round(
                sum(c["eal_total"] * norm_weights[c["fips"]] for c in matched_counties), 2
            ),
        }

        # Re-derive ratings from weighted scores using quintile thresholds
        composite["risk_rating"] = score_to_rating(composite["risk_score"])
        composite["eal_rating"] = score_to_rating(composite["eal_score"])
        composite["sovi_rating"] = score_to_rating(composite["sovi_score"])
        composite["resl_rating"] = score_to_rating(composite["resl_score"])

        # Step 5: Area-weighted per-hazard scores (all 18 types always present)
        all_hazards: dict[str, dict] = {}
        for code in NRI_HAZARD_CODES:
            hazard_records = [
                (c["hazards"][code], norm_weights[c["fips"]])
                for c in matched_counties
                if code in c.get("hazards", {})
            ]
            if not hazard_records:
                # No county data for this hazard type -- include with zeros
                all_hazards[code] = {
                    "risk_score": 0.0,
                    "risk_rating": "",
                    "eal_total": 0.0,
                    "annualized_freq": 0.0,
                    "num_events": 0.0,
                }
                continue

            weighted_risk = sum(h["risk_score"] * w for h, w in hazard_records)
            weighted_eal = sum(h["eal_total"] * w for h, w in hazard_records)
            weighted_freq = sum(h["annualized_freq"] * w for h, w in hazard_records)
            total_events = sum(h["num_events"] * w for h, w in hazard_records)

            all_hazards[code] = {
                "risk_score": round(weighted_risk, 2),
                "risk_rating": score_to_rating(weighted_risk) if weighted_risk > 0 else "",
                "eal_total": round(weighted_eal, 2),
                "annualized_freq": round(weighted_freq, 4),
                "num_events": round(total_events, 2),
            }

        # Step 6: Top 5 hazards by risk score (non-zero only)
        nonzero_hazards = [
            (code, data) for code, data in all_hazards.items()
            if data["risk_score"] > 0
        ]
        sorted_hazards = sorted(
            nonzero_hazards,
            key=lambda x: x[1]["risk_score"],
            reverse=True,
        )
        top_hazards = [
            {
                "type": NRI_HAZARD_CODES[code],
                "code": code,
                "risk_score": data["risk_score"],
                "risk_rating": data["risk_rating"],
                "eal_total": data["eal_total"],
            }
            for code, data in sorted_hazards[:5]
        ]

        return {
            "version": self._nri_version,
            "counties_analyzed": len(matched_counties),
            "composite": composite,
            "top_hazards": top_hazards,
            "all_hazards": all_hazards,
        }

    def _empty_nri_profile(self, note: str) -> dict:
        """Return a minimal NRI profile when no data is available.

        Args:
            note: Explanation of why data is unavailable.

        Returns:
            Dict with null/zero values and a note field.
        """
        # Build all_hazards with zeros for all 18 NRI codes (schema consistency)
        empty_hazards = {
            code: {
                "risk_score": 0.0,
                "risk_rating": "",
                "eal_total": 0.0,
                "annualized_freq": 0.0,
                "num_events": 0.0,
            }
            for code in NRI_HAZARD_CODES
        }
        return {
            "version": "",
            "counties_analyzed": 0,
            "note": note,
            "composite": {
                "risk_score": 0.0,
                "risk_rating": "",
                "eal_total": 0.0,
                "eal_score": 0.0,
                "eal_rating": "",
                "sovi_score": 0.0,
                "sovi_rating": "",
                "resl_score": 0.0,
                "resl_rating": "",
            },
            "top_hazards": [],
            "all_hazards": empty_hazards,
        }

    def _load_usfs_wildfire_data(self) -> dict[str, dict]:
        """Parse the USFS Wildfire Risk to Communities XLSX.

        Searches data/usfs/ for XLSX files and reads wildfire risk metrics.
        Since exact column names are uncertain, performs schema discovery
        by inspecting headers first.

        Returns:
            Dict keyed by geographic identifier (GEOID or name) -> wildfire
            metrics dict. Empty dict if no XLSX found or parse fails.
        """
        if not self.usfs_dir.exists():
            logger.warning(
                "USFS data directory %s does not exist -- "
                "wildfire risk data will be unavailable. "
                "Download from https://wildfirerisk.org/",
                self.usfs_dir,
            )
            return {}

        # Find XLSX files
        xlsx_files = list(self.usfs_dir.glob("*.xlsx"))
        if not xlsx_files:
            logger.warning(
                "No XLSX files found in %s -- wildfire risk data unavailable",
                self.usfs_dir,
            )
            return {}

        xlsx_path = xlsx_files[0]
        if len(xlsx_files) > 1:
            logger.info(
                "Found %d XLSX files in USFS dir, using: %s",
                len(xlsx_files), xlsx_path.name,
            )

        logger.info("Loading USFS wildfire data from %s", xlsx_path)

        try:
            from openpyxl import load_workbook
        except ImportError:
            logger.error(
                "openpyxl not installed -- cannot parse USFS XLSX. "
                "Install with: pip install openpyxl"
            )
            return {}

        try:
            wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        except Exception as exc:
            logger.error(
                "Failed to open USFS XLSX %s: %s", xlsx_path, exc,
            )
            return {}

        ws = wb.active
        if ws is None:
            logger.error("No active sheet in USFS XLSX %s", xlsx_path)
            wb.close()
            return {}

        # Read headers from first row
        rows = ws.iter_rows(min_row=1, max_row=1, values_only=True)
        header_row = next(rows, None)
        if header_row is None:
            logger.error("Empty USFS XLSX %s", xlsx_path)
            wb.close()
            return {}

        headers = [str(h).strip() if h is not None else "" for h in header_row]
        logger.info("USFS XLSX headers: %s", headers)

        # Discover column indices by keyword matching
        geoid_idx = None
        name_idx = None
        risk_idx = None
        likelihood_idx = None

        for i, h in enumerate(headers):
            h_lower = h.lower()
            if geoid_idx is None and any(
                kw in h_lower for kw in ["geoid", "aiannh", "fips"]
            ):
                geoid_idx = i
            if name_idx is None and any(
                kw in h_lower for kw in ["name", "tribal", "community"]
            ):
                name_idx = i
            if risk_idx is None and any(
                kw in h_lower for kw in ["risk_to_homes", "risk score", "whp", "risk"]
            ):
                risk_idx = i
            if likelihood_idx is None and any(
                kw in h_lower
                for kw in ["likelihood", "burn_prob", "burn probability", "probability"]
            ):
                likelihood_idx = i

        # If we found neither geoid nor name, we can't key the data
        if geoid_idx is None and name_idx is None:
            logger.warning(
                "USFS XLSX headers don't contain recognizable GEOID or name columns. "
                "Headers: %s. Wildfire data will be unavailable.",
                headers,
            )
            wb.close()
            return {}

        logger.info(
            "USFS column mapping: geoid_idx=%s, name_idx=%s, risk_idx=%s, likelihood_idx=%s",
            geoid_idx, name_idx, risk_idx, likelihood_idx,
        )

        usfs_data: dict[str, dict] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None:
                continue

            # Determine the key (prefer GEOID, fallback to name)
            key = None
            geoid_val = None
            name_val = None

            if geoid_idx is not None and geoid_idx < len(row) and row[geoid_idx]:
                geoid_val = str(row[geoid_idx]).strip()
                key = geoid_val
            if name_idx is not None and name_idx < len(row) and row[name_idx]:
                name_val = str(row[name_idx]).strip()
                if key is None:
                    key = name_val

            if not key:
                continue

            record = {
                "geoid": geoid_val,
                "name": name_val,
                "risk_to_homes": _safe_float(
                    row[risk_idx] if risk_idx is not None and risk_idx < len(row) else None
                ),
                "wildfire_likelihood": _safe_float(
                    row[likelihood_idx]
                    if likelihood_idx is not None and likelihood_idx < len(row)
                    else None
                ),
            }

            # Capture all numeric columns as extra metrics
            extras = {}
            for i, h in enumerate(headers):
                if i in (geoid_idx, name_idx, risk_idx, likelihood_idx):
                    continue
                if i < len(row) and row[i] is not None:
                    val = _safe_float(row[i], default=None)
                    if val is not None:
                        extras[h] = val
            if extras:
                record["additional_metrics"] = extras

            usfs_data[key] = record

        wb.close()
        logger.info("Loaded USFS wildfire data for %d areas", len(usfs_data))
        return usfs_data

    def _match_usfs_to_tribe(
        self, usfs_data: dict[str, dict], tribe: dict
    ) -> dict | None:
        """Match USFS wildfire data to a specific Tribe.

        Matching strategy:
          1. AIANNH GEOID match (if USFS data is keyed by GEOID)
          2. Name match using normalized string comparison

        Args:
            usfs_data: Full USFS data dict from _load_usfs_wildfire_data().
            tribe: Tribe dict from registry.

        Returns:
            Matched wildfire metrics dict, or None if no match.
        """
        if not usfs_data:
            return None

        tribe_id = tribe["tribe_id"]
        geoids = self._tribe_to_geoids.get(tribe_id, [])

        # Strategy 1: GEOID match
        for geoid in geoids:
            if geoid in usfs_data:
                return usfs_data[geoid]

        # Strategy 2: Name match via normalized comparison
        tribe_name = tribe["name"].lower().strip()
        alt_names = [n.lower().strip() for n in tribe.get("alternate_names", [])]
        all_names = [tribe_name] + alt_names

        for key, record in usfs_data.items():
            record_name = (record.get("name") or key).lower().strip()
            for name in all_names:
                if name == record_name or name in record_name or record_name in name:
                    return record

        return None

    def build_all_profiles(self) -> int:
        """Build and cache hazard profiles for all 592 Tribes.

        Main orchestration method:
          1. Load NRI county data from CSV
          2. Load NRI tribal relational data (AIANNH -> county mapping)
          3. Load USFS wildfire data from XLSX
          4. For each Tribe:
             a. Build NRI profile (area-weighted)
             b. Match USFS wildfire data
             c. Override NRI WFIR with USFS conditional risk (if fire-prone)
             d. Write JSON cache file
          5. Generate coverage report (JSON + Markdown)

        Single-writer-per-tribe_id assumption: this method writes one JSON
        file per tribe_id sequentially. If called concurrently for the same
        tribe_id, the last writer wins. Callers must serialize per-tribe
        writes if concurrent execution is needed.

        Returns:
            Number of profile files written.
        """
        # Ensure cache directory exists
        self.cache_dir.mkdir(parents=True, exist_ok=True)

        # Load data sources
        county_data = self._load_nri_county_data()
        tribal_counties = self._load_nri_tribal_relational()
        usfs_data = self._load_usfs_wildfire_data()

        all_tribes = self._registry.get_all()
        logger.info(
            "Building hazard profiles for %d Tribes "
            "(NRI counties: %d, tribal relations: %d, USFS areas: %d)",
            len(all_tribes),
            len(county_data),
            len(tribal_counties),
            len(usfs_data),
        )

        files_written = 0
        nri_matched = 0
        usfs_matched = 0
        usfs_overrides = 0
        unmatched_tribes: list[dict] = []  # [{name, tribe_id, states}]

        # Per-state tracking for coverage report
        state_stats: dict[str, dict] = {}  # state -> {total, nri_matched, usfs_matched, unmatched, tribe_names}

        for tribe in all_tribes:
            tribe_id = Path(tribe["tribe_id"]).name
            if tribe_id != tribe["tribe_id"] or ".." in tribe_id:
                logger.warning("Skipping suspicious tribe_id: %r", tribe["tribe_id"])
                continue
            tribe_states = tribe.get("states", [])

            # Initialize per-state tracking
            for st in tribe_states:
                if st not in state_stats:
                    state_stats[st] = {
                        "total": 0,
                        "nri_matched": 0,
                        "usfs_matched": 0,
                        "unmatched": 0,
                        "tribe_names": [],
                    }
                state_stats[st]["total"] += 1
                state_stats[st]["tribe_names"].append(tribe["name"])

            # Build NRI profile
            nri_profile = self._build_tribe_nri_profile(
                tribe_id, tribe, county_data, tribal_counties,
            )
            has_nri = nri_profile.get("counties_analyzed", 0) > 0
            if has_nri:
                nri_matched += 1
                for st in tribe_states:
                    if st in state_stats:
                        state_stats[st]["nri_matched"] += 1

            # Match USFS wildfire data
            usfs_match = self._match_usfs_to_tribe(usfs_data, tribe)
            usfs_profile: dict = {}
            if usfs_match:
                usfs_matched += 1
                usfs_profile = {
                    "risk_to_homes": usfs_match.get("risk_to_homes", 0.0),
                    "conditional_risk_to_structures": usfs_match.get("risk_to_homes", 0.0),
                    "wildfire_likelihood": usfs_match.get("wildfire_likelihood", 0.0),
                    "source_name": usfs_match.get("name", ""),
                }
                if "additional_metrics" in usfs_match:
                    usfs_profile["additional_metrics"] = usfs_match["additional_metrics"]
                for st in tribe_states:
                    if st in state_stats:
                        state_stats[st]["usfs_matched"] += 1

            # USFS wildfire override: replace NRI WFIR when USFS data exists
            # and Tribe is fire-prone (NRI WFIR score > 0)
            wfir_entry = nri_profile.get("all_hazards", {}).get("WFIR")
            if usfs_match and wfir_entry and wfir_entry.get("risk_score", 0.0) > 0:
                nri_wfir_original = wfir_entry.get("risk_score", 0.0)

                # Override with USFS conditional risk to structures
                usfs_risk = usfs_match.get("risk_to_homes", 0.0)
                if usfs_risk > 0:
                    wfir_entry["risk_score"] = usfs_risk
                    wfir_entry["risk_rating"] = score_to_rating(usfs_risk)
                    wfir_entry["source"] = "USFS"  # Mark the data source

                    # Store original NRI value in usfs_wildfire section
                    usfs_profile["nri_wfir_original"] = round(nri_wfir_original, 2)
                    usfs_overrides += 1

                    # Re-extract top 5 after USFS override (non-zero only)
                    nonzero = [
                        (code, data) for code, data in nri_profile["all_hazards"].items()
                        if data["risk_score"] > 0
                    ]
                    sorted_hazards = sorted(
                        nonzero,
                        key=lambda x: x[1]["risk_score"],
                        reverse=True,
                    )
                    nri_profile["top_hazards"] = [
                        {
                            "type": NRI_HAZARD_CODES.get(code, code),
                            "code": code,
                            "risk_score": data["risk_score"],
                            "risk_rating": data["risk_rating"],
                            "eal_total": data["eal_total"],
                        }
                        for code, data in sorted_hazards[:5]
                    ]

            # Track unmatched
            if not has_nri and not usfs_match:
                unmatched_tribes.append({
                    "name": tribe["name"],
                    "tribe_id": tribe_id,
                    "states": tribe_states,
                })
                for st in tribe_states:
                    if st in state_stats:
                        state_stats[st]["unmatched"] += 1

            # Combine into full profile
            profile = {
                "tribe_id": tribe_id,
                "tribe_name": tribe["name"],
                "sources": {
                    "fema_nri": nri_profile,
                    "usfs_wildfire": usfs_profile,
                },
                "generated_at": datetime.now(timezone.utc).isoformat(),
            }

            # Write cache file atomically
            cache_file = self.cache_dir / f"{tribe_id}.json"
            atomic_write_json(cache_file, profile)
            files_written += 1

        # Log summary
        logger.info(
            "Hazard profile build complete: %d files written, "
            "%d with NRI data, %d with USFS data, %d USFS overrides, "
            "%d fully unmatched",
            files_written,
            nri_matched,
            usfs_matched,
            usfs_overrides,
            len(unmatched_tribes),
        )

        if unmatched_tribes:
            names = [t["name"] for t in unmatched_tribes]
            logger.warning(
                "%d Tribes not covered by federal hazard datasets "
                "(federal data gap — no NRI county crosswalk or USFS match): %s%s",
                len(names),
                ", ".join(names[:10]),
                f"... and {len(names) - 10} more"
                if len(names) > 10 else "",
            )

        # Generate coverage report
        coverage_data = {
            "total_profiles": files_written,
            "nri_matched": nri_matched,
            "usfs_matched": usfs_matched,
            "usfs_overrides": usfs_overrides,
            "scored_profiles": nri_matched,  # profiles with counties_analyzed > 0
            "unmatched_count": len(unmatched_tribes),
            "unmatched_tribes": unmatched_tribes,
            "state_stats": state_stats,
        }
        self._generate_coverage_report(coverage_data)

        return files_written

    def _generate_coverage_report(self, coverage_data: dict) -> None:
        """Generate hazard coverage report in JSON and Markdown formats.

        Writes two files to the outputs directory:
          - hazard_coverage_report.json (machine-readable)
          - hazard_coverage_report.md (human-readable with tables)

        Args:
            coverage_data: Dict with tracking data accumulated during
                build_all_profiles(), including counts, unmatched list,
                and per-state statistics.
        """
        total = coverage_data["total_profiles"]
        nri = coverage_data["nri_matched"]
        usfs = coverage_data["usfs_matched"]
        usfs_overrides = coverage_data["usfs_overrides"]
        scored = coverage_data["scored_profiles"]
        unmatched_count = coverage_data["unmatched_count"]
        unmatched_tribes = coverage_data["unmatched_tribes"]
        state_stats = coverage_data["state_stats"]

        # Build JSON report
        report_json = {
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "summary": {
                "total_profiles": total,
                "nri_matched": nri,
                "usfs_matched": usfs,
                "usfs_overrides": usfs_overrides,
                "scored_profiles": scored,
                "unmatched_profiles": unmatched_count,
                "scored_pct": round(scored / max(total, 1) * 100, 1),
                "nri_pct": round(nri / max(total, 1) * 100, 1),
                "usfs_pct": round(usfs / max(total, 1) * 100, 1),
            },
            "unmatched_tribes": [
                {"name": t["name"], "tribe_id": t["tribe_id"], "states": t["states"]}
                for t in unmatched_tribes
            ],
            "per_state": {
                st: {
                    "total": stats["total"],
                    "nri_matched": stats["nri_matched"],
                    "usfs_matched": stats["usfs_matched"],
                    "unmatched": stats["unmatched"],
                }
                for st, stats in sorted(state_stats.items())
            },
        }

        json_path = OUTPUTS_DIR / "hazard_coverage_report.json"
        atomic_write_json(json_path, report_json)
        logger.info("Coverage report (JSON) written to %s", json_path)

        # Build Markdown report
        md_lines = [
            "# Hazard Coverage Report",
            "",
            f"Generated: {report_json['generated_at']}",
            "",
            "## Summary",
            "",
            "| Metric | Value |",
            "|--------|-------|",
            f"| Total profiles | {total} |",
            f"| NRI matched | {nri} ({report_json['summary']['nri_pct']}%) |",
            f"| USFS matched | {usfs} ({report_json['summary']['usfs_pct']}%) |",
            f"| USFS overrides (WFIR) | {usfs_overrides} |",
            f"| Scored profiles | {scored} ({report_json['summary']['scored_pct']}%) |",
            f"| Unmatched | {unmatched_count} |",
            "",
            "## Per-State Breakdown",
            "",
            "| State | Total Tribes | NRI Matched | USFS Matched | Unmatched |",
            "|-------|-------------|-------------|--------------|-----------|",
        ]

        for st in sorted(state_stats.keys()):
            stats = state_stats[st]
            md_lines.append(
                f"| {st} | {stats['total']} | {stats['nri_matched']} "
                f"| {stats['usfs_matched']} | {stats['unmatched']} |"
            )

        md_lines.extend([
            "",
            "## Tribes Not Covered by Federal Hazard Data",
            "",
        ])

        if unmatched_tribes:
            for t in unmatched_tribes:
                states_str = ", ".join(t["states"]) if t["states"] else "N/A"
                md_lines.append(f"- {t['name']} ({t['tribe_id']}) -- {states_str}")
        else:
            md_lines.append("None -- all Tribes have hazard data.")

        md_lines.append("")  # trailing newline

        md_path = OUTPUTS_DIR / "hazard_coverage_report.md"
        atomic_write_text(md_path, "\n".join(md_lines))
        logger.info("Coverage report (Markdown) written to %s", md_path)
